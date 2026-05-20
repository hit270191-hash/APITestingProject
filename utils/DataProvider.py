import json
import csv
from openpyxl import load_workbook
import os

#Json Data Provider
def read_json_data(filename):
    with open(filename, 'r') as f:
        data = json.load(f)
    return [(item,) for item in data]


def read_excel_data(filename, sheet_name):
    wb =load_workbook(filename)
    sheet = wb[sheet_name]

    header= [cell.value for cell in sheet[1]]
    data=[]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row):
            data.append(dict(zip(header,row)))
    return data

# xlpath= os.path.abspath(os.path.join(os.path.dirname(__file__),"../testData/products_data.xlsx"))
# read_excel_data(xlpath, "products_data")